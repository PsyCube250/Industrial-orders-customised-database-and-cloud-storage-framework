from datetime import datetime, date, timedelta
from statistics import mean

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from openpyxl import load_workbook
from sqlalchemy import func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload, selectinload

from app.auth import get_current_user
from app.database import get_db
from app.models.notification import TimeoutRule
from app.models.order import Order, OrderAttachment, OrderProgressStep, OrderChangeLog, ORDER_WORKFLOW_STEPS
from app.models.system import OperationLog
from app.uploads import save_upload
from app.schemas.order import (
    OrderCreate,
    OrderUpdate,
    OrderChangeAction,
    OrderOut,
    OrderDetailOut,
    OrderStatsOut,
)

router = APIRouter(prefix="/api/orders", tags=["orders"])


def _next_order_no(db: Session) -> str:
    today_prefix = f"ORD{datetime.utcnow():%Y%m%d}"
    last = (
        db.query(Order.order_no)
        .filter(Order.order_no.like(f"{today_prefix}%"))
        # length before value: once the sequence outgrows its 4-digit padding,
        # plain string ordering would put e.g. 9999 above 10000
        .order_by(func.length(Order.order_no).desc(), Order.order_no.desc())
        .first()
    )
    seq = int(last[0][len(today_prefix):]) + 1 if last else 1
    return f"{today_prefix}{seq:04d}"


def _step_due_at(db: Session, step_name: str, started_at: datetime) -> datetime | None:
    rule = db.query(TimeoutRule).filter(TimeoutRule.step_name == step_name).first()
    return started_at + timedelta(days=rule.days_allowed) if rule else None


def _seed_progress_steps(db: Session, order: Order) -> list[OrderProgressStep]:
    steps = []
    now = datetime.utcnow()
    for i, name in enumerate(ORDER_WORKFLOW_STEPS):
        steps.append(
            OrderProgressStep(
                step_name=name,
                sequence=i,
                status="in_progress" if i == 0 else "pending",
                started_at=now if i == 0 else None,
                due_at=_step_due_at(db, name, now) if i == 0 else None,
            )
        )
    return steps


def _create_order_with_steps(db: Session, **fields) -> Order:
    """Insert an order inside a savepoint, retrying if a concurrent request
    grabbed the same order number (unique constraint on order_no)."""
    last_exc = None
    for _ in range(3):
        try:
            with db.begin_nested():
                order = Order(order_no=_next_order_no(db), **fields)
                db.add(order)
                db.flush()
                for step in _seed_progress_steps(db, order):
                    step.order_id = order.id
                    db.add(step)
            return order
        except IntegrityError as exc:
            last_exc = exc
    raise HTTPException(status_code=409, detail="Could not allocate an order number, please retry") from last_exc


def _log(db: Session, user, action: str, target: str):
    db.add(OperationLog(user_id=getattr(user, "id", None), action=action, target=target))


@router.post("", response_model=OrderDetailOut)
def create_order(
    payload: OrderCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)
):
    order = _create_order_with_steps(
        db,
        customer_name=payload.customer_name,
        product_name=payload.product_name,
        quantity=payload.quantity,
        delivery_date=payload.delivery_date,
        notes=payload.notes,
    )
    _log(db, current_user, "create_order", order.order_no)
    db.commit()
    db.refresh(order)
    return order


@router.post("/bulk-import")
def bulk_import_orders(
    file: UploadFile = File(...), db: Session = Depends(get_db), current_user=Depends(get_current_user)
):
    """Import orders from an Excel template with columns:
    customer_name | product_name | quantity | delivery_date (YYYY-MM-DD) | notes
    """
    try:
        wb = load_workbook(file.file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail="Not a valid .xlsx file") from exc
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        if row is None or all(c is None for c in row):
            continue
        try:
            # savepoint per row: a failed insert rolls back just this row instead of
            # poisoning the session for the rest of the import
            with db.begin_nested():
                customer_name, product_name, quantity, delivery_date, notes = (list(row) + [None] * 5)[:5]
                if not customer_name:
                    raise ValueError("customer_name is required")
                quantity = int(quantity or 0)
                if quantity < 0:
                    raise ValueError("quantity must be >= 0")
                if isinstance(delivery_date, str) and delivery_date:
                    delivery_date = datetime.strptime(delivery_date, "%Y-%m-%d").date()
                elif isinstance(delivery_date, datetime):
                    delivery_date = delivery_date.date()
                order = _create_order_with_steps(
                    db,
                    customer_name=str(customer_name),
                    product_name=str(product_name or ""),
                    quantity=quantity,
                    delivery_date=delivery_date if isinstance(delivery_date, date) else None,
                    notes=str(notes or ""),
                )
            created.append(order.order_no)
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": idx, "error": str(exc)})
    _log(db, current_user, "bulk_import_orders", f"{len(created)} orders")
    db.commit()
    return {"created": created, "errors": errors}


@router.get("", response_model=list[OrderOut])
def search_orders(
    keyword: str | None = Query(None, description="matches customer name or order number"),
    status: str | None = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(Order)
    if keyword:
        like = f"%{keyword}%"
        q = q.filter(or_(Order.customer_name.ilike(like), Order.order_no.ilike(like)))
    if status:
        q = q.filter(Order.status == status)
    return q.order_by(Order.created_at.desc()).all()


@router.get("/stats", response_model=OrderStatsOut)
def order_stats(db: Session = Depends(get_db), _=Depends(get_current_user)):
    orders = db.query(Order).options(joinedload(Order.progress_steps)).all()
    total = len(orders)

    def completion_date(o: Order) -> date:
        # last finished step marks completion; updated_at moves on any later edit
        done = [s.completed_at for s in o.progress_steps if s.completed_at]
        return (max(done) if done else o.updated_at).date()

    today = datetime.utcnow().date()
    completed = [o for o in orders if o.status == "completed"]
    on_time = [
        o for o in completed if o.delivery_date and completion_date(o) <= o.delivery_date
    ]
    on_time_rate = round(len(on_time) / len(completed) * 100, 1) if completed else 0.0

    delay_ranking = []
    for o in orders:
        if o.delivery_date:
            reference = completion_date(o) if o.status == "completed" else today
            delay_days = (reference - o.delivery_date).days
            if delay_days > 0:
                delay_ranking.append({"order_no": o.order_no, "customer_name": o.customer_name, "delay_days": delay_days})
    delay_ranking.sort(key=lambda x: x["delay_days"], reverse=True)

    step_durations: dict[str, list[float]] = {}
    for step in db.query(OrderProgressStep).filter(OrderProgressStep.completed_at.isnot(None)).all():
        if step.started_at:
            hours = (step.completed_at - step.started_at).total_seconds() / 3600
            step_durations.setdefault(step.step_name, []).append(hours)
    step_duration_avg = [
        {"step_name": name, "avg_hours": round(mean(vals), 1)} for name, vals in step_durations.items()
    ]

    product_totals: dict[str, int] = {}
    for o in orders:
        key = o.product_name or "(unspecified)"
        product_totals[key] = product_totals.get(key, 0) + o.quantity
    product_quantity_totals = [
        {"product_name": name, "total_quantity": qty} for name, qty in product_totals.items()
    ]

    return OrderStatsOut(
        total_orders=total,
        on_time_rate=on_time_rate,
        delay_ranking=delay_ranking,
        step_duration_avg_hours=step_duration_avg,
        product_quantity_totals=product_quantity_totals,
    )


@router.get("/{order_id}", response_model=OrderDetailOut)
def get_order(order_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    order = (
        db.query(Order)
        .options(
            # selectin, not joined: three joined collections would multiply into a
            # cartesian product (attachments × steps × change logs) per order row
            selectinload(Order.attachments),
            selectinload(Order.progress_steps),
            selectinload(Order.change_logs),
        )
        .filter(Order.id == order_id)
        .first()
    )
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


@router.patch("/{order_id}", response_model=OrderDetailOut)
def update_order(
    order_id: int, payload: OrderUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)
):
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(order, field, value)
    _log(db, current_user, "update_order", order.order_no)
    db.commit()
    db.refresh(order)
    return order


@router.post("/{order_id}/change", response_model=OrderDetailOut)
def change_order(
    order_id: int, action: OrderChangeAction, db: Session = Depends(get_db), current_user=Depends(get_current_user)
):
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if action.change_type == "delivery_date":
        if not action.new_delivery_date:
            raise HTTPException(status_code=400, detail="new_delivery_date is required")
        old_value = str(order.delivery_date) if order.delivery_date else ""
        order.delivery_date = action.new_delivery_date
        new_value = str(order.delivery_date)
    elif action.change_type == "cancel":
        if order.status in ("cancelled", "completed"):
            raise HTTPException(status_code=400, detail=f"Cannot cancel a {order.status} order")
        old_value, order.status, new_value = order.status, "cancelled", "cancelled"
    elif action.change_type == "pause":
        if order.status != "in_progress":
            raise HTTPException(status_code=400, detail="Can only pause an in-progress order")
        old_value, order.status, new_value = order.status, "paused", "paused"
    elif action.change_type == "resume":
        if order.status != "paused":
            raise HTTPException(status_code=400, detail="Can only resume a paused order")
        old_value, order.status, new_value = order.status, "in_progress", "in_progress"
    else:
        raise HTTPException(status_code=400, detail="Unknown change_type")

    db.add(
        OrderChangeLog(
            order_id=order.id,
            change_type=action.change_type,
            old_value=old_value,
            new_value=new_value,
            changed_by=current_user.id,
        )
    )
    _log(db, current_user, f"order_change:{action.change_type}", order.order_no)
    db.commit()
    db.refresh(order)
    return order


@router.post("/{order_id}/progress/{step_id}/advance", response_model=OrderDetailOut)
def advance_progress_step(
    order_id: int, step_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)
):
    """Mark a step done and start the next pending step, driving the tracking progress bar."""
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.status != "in_progress":
        raise HTTPException(status_code=400, detail=f"Order is {order.status}; steps can only advance on an in-progress order")
    step = db.get(OrderProgressStep, step_id)
    if not step or step.order_id != order_id:
        raise HTTPException(status_code=404, detail="Step not found")
    if step.status != "in_progress":
        # only the active step can advance — keeps steps in sequence and stops a
        # done step from being re-completed (which would reset the next step's clock)
        raise HTTPException(status_code=400, detail=f"Step is {step.status}; only the in-progress step can advance")

    now = datetime.utcnow()
    step.status = "done"
    step.completed_at = now

    next_step = (
        db.query(OrderProgressStep)
        .filter(OrderProgressStep.order_id == order_id, OrderProgressStep.sequence == step.sequence + 1)
        .first()
    )
    if next_step:
        next_step.status = "in_progress"
        next_step.started_at = now
        next_step.due_at = _step_due_at(db, next_step.step_name, now)
    else:
        order.status = "completed"

    _log(db, current_user, "advance_progress_step", f"{order.order_no}:{step.step_name}")
    db.commit()
    db.refresh(order)
    return order


@router.post("/{order_id}/attachments", response_model=OrderDetailOut)
def upload_attachment(
    order_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    file_name, dest_path = save_upload(file, "orders", str(order_id))
    db.add(OrderAttachment(order_id=order_id, file_name=file_name, file_path=dest_path))
    _log(db, current_user, "upload_attachment", f"{order.order_no}:{file_name}")
    db.commit()
    db.refresh(order)
    return order
